from __future__ import annotations

import json
from pathlib import Path

import requests
from openpyxl import load_workbook

SOURCES = {
    "2020_receipts": "https://www.soumu.go.jp/main_sosiki/jichi_zeisei/czaisei/czaisei_seido/furusato/file/results20210730-01.xlsx",
    "2020_tax": "https://www.soumu.go.jp/main_sosiki/jichi_zeisei/czaisei/czaisei_seido/furusato/file/results20210730-03.xlsx",
    "2021_receipts": "https://www.soumu.go.jp/main_sosiki/jichi_zeisei/czaisei/czaisei_seido/furusato/file/results20220729-01.xlsx",
    "2021_tax": "https://www.soumu.go.jp/main_sosiki/jichi_zeisei/czaisei/czaisei_seido/furusato/file/results20220729-03.xlsx",
    "2022_receipts": "https://www.soumu.go.jp/main_content/000894504.xlsx",
    "2022_tax": "https://www.soumu.go.jp/main_content/000897134.xlsx",
    "2023_receipts": "https://www.soumu.go.jp/main_content/000960672.xlsx",
    "2023_tax": "https://www.soumu.go.jp/main_content/000960675.xlsx",
    "2024_receipts": "https://www.soumu.go.jp/main_content/001022818.xlsx",
    "2024_tax": "https://www.soumu.go.jp/main_content/001022820.xlsx",
}


def clean(value):
    if value is None:
        return None
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return text[:180]


def inspect_book(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"sheetnames": wb.sheetnames, "sheets": {}}
    for ws in wb.worksheets:
        rows = []
        max_col = min(ws.max_column or 0, 80)
        max_row = min(ws.max_row or 0, 28)
        for r in range(1, max_row + 1):
            vals = [clean(ws.cell(r, c).value) for c in range(1, max_col + 1)]
            if any(v not in (None, "") for v in vals):
                rows.append({"row": r, "values": vals})
        result["sheets"][ws.title] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample_rows": rows,
        }
    return result


def main():
    out = Path("tmp_build/output")
    raw = out / "raw"
    raw.mkdir(parents=True, exist_ok=True)
    report = {}
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 source-audit"})
    for key, url in SOURCES.items():
        entry = {"url": url}
        try:
            response = session.get(url, timeout=120)
            entry["status"] = response.status_code
            entry["content_type"] = response.headers.get("content-type")
            entry["size"] = len(response.content)
            response.raise_for_status()
            path = raw / f"{key}.xlsx"
            path.write_bytes(response.content)
            entry["workbook"] = inspect_book(path)
        except Exception as exc:
            entry["error"] = repr(exc)
        report[key] = entry
    (out / "diagnostics.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: {kk: vv for kk, vv in v.items() if kk != "workbook"} for k, v in report.items()}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
